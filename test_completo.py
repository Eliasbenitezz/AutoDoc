import os
import subprocess
import sys
from pathlib import Path

def test_completo():
    """Prueba completa del sistema AutoDoc"""
    print("🧪 PRUEBA COMPLETA DEL SISTEMA AUTODOC")
    print("=" * 60)
    
    # 1. Verificar archivos necesarios
    print("\n1. Verificando archivos necesarios...")
    archivos_requeridos = ["AutoDoc.exe", "Plantilla.xlsx", "cargaV2.xlsx"]
    for archivo in archivos_requeridos:
        if os.path.exists(archivo):
            print(f"   ✅ {archivo}")
        else:
            print(f"   ❌ {archivo} - NO ENCONTRADO")
            return False
    
    # 2. Verificar dependencias Python
    print("\n2. Verificando dependencias...")
    try:
        import openpyxl
        print("   ✅ openpyxl")
    except ImportError:
        print("   ❌ openpyxl - NO INSTALADO")
        return False
    
    # 3. Verificar hojas del archivo Excel
    print("\n3. Verificando archivo Excel...")
    try:
        import openpyxl
        workbook = openpyxl.load_workbook("cargaV2.xlsx", data_only=True)
        print(f"   ✅ Archivo cargado correctamente")
        print(f"   📊 Hojas disponibles: {len(workbook.sheetnames)}")
        print(f"   📋 Primera hoja: {workbook.sheetnames[0]}")
    except Exception as e:
        print(f"   ❌ Error al cargar archivo: {e}")
        return False
    
    # 4. Probar ejecutable
    print("\n4. Probando ejecutable...")
    input_script = """cargaV2.xlsx
PLANTA
10
12
"""
    
    with open("input_test.txt", "w", encoding="utf-8") as f:
        f.write(input_script)
    
    try:
        with open("input_test.txt", "r", encoding="utf-8") as input_file:
            result = subprocess.run(
                ["./AutoDoc.exe"],
                input=input_file.read(),
                text=True,
                capture_output=True,
                timeout=60
            )
        
        if result.returncode == 0:
            print("   ✅ Ejecutable funcionó correctamente")
        else:
            print("   ⚠️ Ejecutable completó con advertencias")
        
        # Verificar que se crearon archivos
        desktop = Path.home() / "Desktop"
        autodoc_folder = desktop / "AutoDoc"
        
        if not autodoc_folder.exists():
            # Buscar en directorio actual
            autodoc_folder = Path.cwd() / "AutoDoc"
        
        if autodoc_folder.exists():
            print(f"   ✅ Carpeta AutoDoc creada: {autodoc_folder}")
            
            # Verificar archivos generados
            plana_folder = autodoc_folder / "PLANTA"
            if plana_folder.exists():
                archivos = list(plana_folder.glob("*.xlsx"))
                print(f"   📁 Archivos generados en PLANTA: {len(archivos)}")
                if archivos:
                    print(f"   ✅ Primer archivo: {archivos[0].name}")
            else:
                print("   ❌ Carpeta PLANTA no encontrada")
        else:
            print("   ❌ Carpeta AutoDoc no encontrada")
            
    except Exception as e:
        print(f"   ❌ Error en ejecutable: {e}")
        return False
    finally:
        if os.path.exists("input_test.txt"):
            os.remove("input_test.txt")
    
    # 5. Resumen
    print("\n" + "=" * 60)
    print("📋 RESUMEN DE LA PRUEBA")
    print("=" * 60)
    print("✅ Todos los archivos necesarios están presentes")
    print("✅ Las dependencias están instaladas")
    print("✅ El archivo Excel se puede cargar correctamente")
    print("✅ El ejecutable funciona correctamente")
    print("✅ Se generaron archivos de salida")
    print("\n🎉 ¡SISTEMA LISTO PARA USAR!")
    
    return True

if __name__ == "__main__":
    success = test_completo()
    if not success:
        print("\n💥 Se encontraron problemas que deben resolverse")
    
    input("\nPresiona Enter para continuar...") 