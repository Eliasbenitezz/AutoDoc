import warnings
warnings.filterwarnings("ignore", category=UserWarning)
import openpyxl
import random
import os
import shutil
import sys
from pathlib import Path

# Configurar codificación para Windows
if sys.platform.startswith('win'):
    import locale
    if locale.getpreferredencoding().upper() != 'UTF-8':
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')

def crear_carpeta_autodoc():
    """Crea la carpeta AutoDoc en el escritorio y copia la plantilla"""
    # Intentar diferentes ubicaciones del escritorio
    possible_desktops = [
        Path.home() / "Desktop",
        Path.home() / "Escritorio",
        Path.home() / "Área de trabajo",
        Path.home() / "Desktop (C:)",
    ]
    
    desktop = None
    for possible_desktop in possible_desktops:
        if possible_desktop.exists():
            desktop = possible_desktop
            break
    
    # Si no se encuentra el escritorio, usar el directorio actual
    if desktop is None:
        desktop = Path.cwd()
        print(f"ADVERTENCIA: No se pudo encontrar el escritorio. Usando directorio actual: {desktop}")
    
    autodoc_path = desktop / "AutoDoc"
    
    # Crear carpeta AutoDoc si no existe
    try:
        if not autodoc_path.exists():
            autodoc_path.mkdir(parents=True, exist_ok=True)
            print(f"Carpeta 'AutoDoc' creada en: {autodoc_path}")
        else:
            print(f"Carpeta 'AutoDoc' ya existe en: {autodoc_path}")
    except Exception as e:
        print(f"Error al crear carpeta: {e}")
        # Fallback: usar directorio actual
        autodoc_path = Path.cwd() / "AutoDoc"
        autodoc_path.mkdir(parents=True, exist_ok=True)
        print(f"Carpeta 'AutoDoc' creada en directorio actual: {autodoc_path}")
    
    # Copiar plantilla.xlsx a la carpeta AutoDoc
    plantilla_origen = Path("Plantilla.xlsx")
    plantilla_destino = autodoc_path / "Plantilla.xlsx"
    
    if plantilla_origen.exists():
        shutil.copy2(plantilla_origen, plantilla_destino)
        print(f"Plantilla.xlsx copiada a: {plantilla_destino}")
    else:
        print("Error: No se encontró Plantilla.xlsx en el directorio actual")
        return None
    
    return autodoc_path

def solicitar_parametros():
    """Solicita los parámetros al usuario"""
    print("\n" + "="*60)
    print("           CONFIGURACION DE AUTODOC")
    print("="*60)
    
    # 1. Nombre del archivo xlsx
    while True:
        input_filename = input("\n1. Ingrese el nombre del archivo Excel (ej: datos.xlsx): ").strip()
        if input_filename:
            if not input_filename.endswith('.xlsx'):
                input_filename += '.xlsx'
            break
        print("Error: Debe ingresar un nombre de archivo valido.")
    
    # 2. Nombres de las hojas
    print("\n2. Ingrese los nombres de las hojas a procesar:")
    print("   (Puede ingresar multiples hojas separadas por comas)")
    sheet_input_names = []
    while True:
        sheets_input = input("   Hojas: ").strip()
        if sheets_input:
            sheet_input_names = [sheet.strip() for sheet in sheets_input.split(',')]
            break
        print("Error: Debe ingresar al menos una hoja.")
    
    # 3. Rango de filas para cada hoja
    print("\n3. Configuración de rangos de filas:")
    rangos_hojas = {}
    
    for sheet_name in sheet_input_names:
        print(f"\n   Para la hoja '{sheet_name}':")
        while True:
            try:
                fila_inicio = int(input(f"   Fila de inicio: "))
                fila_fin = int(input(f"   Fila de fin: "))
                if fila_inicio <= fila_fin and fila_inicio > 0:
                    rangos_hojas[sheet_name] = (fila_inicio, fila_fin)
                    break
                else:
                    print("Error: La fila de inicio debe ser menor o igual a la fila de fin, y ambas deben ser positivas.")
            except ValueError:
                print("Error: Debe ingresar numeros validos.")
    
    return input_filename, sheet_input_names, rangos_hojas

def procesar_archivo(input_filename, sheet_input_names, rangos_hojas, autodoc_path):
    """Procesa el archivo Excel según los parámetros especificados"""
    
    # Guardar el directorio original
    original_dir = os.getcwd()
    
    # Copiar el archivo de entrada a la carpeta AutoDoc si no está ahí
    input_path = Path(input_filename)
    if not input_path.is_absolute():
        input_path = Path(original_dir) / input_filename
    
    dest_path = autodoc_path / input_filename
    if not dest_path.exists():
        try:
            shutil.copy2(input_path, dest_path)
            print(f"Archivo '{input_filename}' copiado a la carpeta AutoDoc.")
        except Exception as e:
            print(f"Error al copiar el archivo: {e}")
            return
    
    # Cambiar al directorio AutoDoc
    os.chdir(autodoc_path)
    plantilla_filename = 'Plantilla.xlsx'
    
    try:
        workbook_input = openpyxl.load_workbook(input_filename, data_only=True) 
        print(f"\nArchivo '{input_filename}' cargado correctamente.")
    except FileNotFoundError:
        print(f"\nError: El archivo '{input_filename}' no se encontró.")
        print("Asegúrate de que el archivo esté en la carpeta AutoDoc o proporciona la ruta completa.")
        return
    except Exception as e:
        print(f"\nError al cargar el archivo: {e}")
        return

    for sheet_name in sheet_input_names:
        fila_inicio, fila_fin = rangos_hojas[sheet_name]
        
        print(f"\n{'='*50}")
        print(f"Procesando hoja: {sheet_name}")
        print(f"Rango de filas: {fila_inicio} - {fila_fin}")
        print(f"{'='*50}")
        
        if sheet_name not in workbook_input.sheetnames:
            print(f"Error: La hoja '{sheet_name}' no existe en el archivo '{input_filename}'.")
            print(f"Saltando a la siguiente hoja...")
            continue
        
        sheet_input = workbook_input[sheet_name]
        print(f"Hoja '{sheet_name}' seleccionada correctamente.")
        
        output_dir = sheet_name
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            print(f"Carpeta '{output_dir}' creada.")
        
        archivos_procesados = 0
        
        # Procesar solo el rango de filas especificado
        for row_index in range(fila_inicio, fila_fin + 1):
            # Verificar que la fila existe en la hoja
            if row_index > sheet_input.max_row:
                print(f"Advertencia: La fila {row_index} no existe en la hoja. Saltando...")
                continue
                
            row = sheet_input[row_index]
            fecha = row[0].value
            modelo = row[1].value
            nrchasis = row[3].value 
            nrmotor = row[4].value 

            if modelo and nrchasis:
                print(f"Procesando fila {row_index}: {modelo} {nrchasis}")

                try:
                    workbook_output = openpyxl.load_workbook(plantilla_filename)
                    sheet_output = workbook_output.active
                except FileNotFoundError:
                    print(f"Error: La plantilla '{plantilla_filename}' no se encontró.")
                    return

                sheet_output['B2'] = fecha
                sheet_output['N2'] = nrchasis 
                sheet_output['E3'] = nrchasis
                sheet_output['E5'] = nrchasis
                sheet_output['L5'] = nrmotor
                sheet_output['S4'] = modelo
                sheet_output['D22'] = random.randint(30, 40)
                sheet_output['D23'] = random.randint(90,120)
                sheet_output['D24'] = random.randint(5,10)
                sheet_output['D25'] = random.randint(5,10)
                sheet_output['D26'] = random.randint(0,3)
                sheet_output['D27'] = random.randint(0,3)
                sheet_output['D28'] = random.randint(80,100)
                sheet_output['M12'] = random.randint(70,90) 
                sheet_output['G18'] = random.randint(250,260)
                sheet_output['J18'] = random.randint(180,200)
                sheet_output['N18'] = random.randint(100,140)
                sheet_output['Q18'] = random.randint(100,140) 

                safe_nombre = str(nrchasis)
                output_filename = f"{nrchasis}.xlsx"
                output_filepath = os.path.join(output_dir, output_filename)

                try:
                    workbook_output.save(output_filepath)
                    print(f"Archivo '{output_filepath}' creado y guardado para {nrchasis}.")
                    archivos_procesados += 1
                except Exception as e:
                    print(f"Error al guardar el archivo '{output_filepath}': {e}")
            else:
                print(f"Advertencia: Fila {row_index} omitida. Nrchasis faltante.")
        
        print(f"\nHoja '{sheet_name}' completada. Archivos procesados: {archivos_procesados}")

    print(f"\n{'='*50}")
    print("--- Proceso completado para todas las hojas ---")
    print(f"{'='*50}")

def main():
    """Función principal del programa"""
    print("AUTODOC - Generador de Documentos Automatico")
    print("="*50)
    
    # Crear carpeta AutoDoc en el escritorio
    autodoc_path = crear_carpeta_autodoc()
    if not autodoc_path:
        try:
            input("\nPresione Enter para salir...")
        except EOFError:
            pass
        return
    
    # Solicitar parámetros
    input_filename, sheet_input_names, rangos_hojas = solicitar_parametros()
    
    # Procesar archivo
    procesar_archivo(input_filename, sheet_input_names, rangos_hojas, autodoc_path)
    
    print(f"\nProceso completado. Los archivos generados se encuentran en:")
    print(f"{autodoc_path}")
    try:
        input("\nPresione Enter para salir...")
    except EOFError:
        pass

if __name__ == "__main__":
    main() 